import tabula
import pandas as pd
import openpyxl
import random
import logging

def convert_pdf_to_excel(path, pdf_to_excel_path):
    print('------Convert PDF to EXCEL------')
    logging.warning('------Convert PDF to EXCEL------')
    # Columns iterpreted as str
    df = pd.DataFrame()
    col2str = {'dtype': str}
    kwargs = {'pages': 'all',
              'pandas_options': col2str,
              'stream': True,
              'lattice': True,
              'silent': True
            }
    all_tables = tabula.read_pdf(path, **kwargs)
    for i in all_tables:    #x values in list []        
        df_table = pd.DataFrame(i).replace({r'\\r': ''}, regex=True)
        df = df.append(df_table)
        df = df.drop(columns=['Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4', 'Unnamed: 5', 'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9', 'Unnamed: 10', 'Unnamed: 11', 'Unnamed: 12'])
    df.to_excel(pdf_to_excel_path, header=True, index = True)
    print('------Finish convert------')
    logging.warning('------Finish convert------')

def get_cid_from_pdf(pdf_to_excel_path):
    print('------Start get company id from PDF------')
    logging.warning('------Start get company id from PDF------')
    companies_id = []
    wb_obj = openpyxl.load_workbook(pdf_to_excel_path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    print(max_row)
    for i in range(2,max_row+1,1):
        # rnum = random.randint(2, max_row)
        cell_obj = sheet_obj.cell(row = i, column = 3)
        # print('Row: ' + str(rnum) + ' Value is: ' + cell_obj.value)
        # logging.info('Row: ' + str(rnum) + ' Value is: ' + cell_obj.value)
        companies_id.append(cell_obj.value)
    # print(sheet_obj.cell(row = max_row, column = 3).value)      
    return companies_id

# def get_cid_from_pdf(pdf_to_excel_path):
#     print('------Start get company id from PDF------')
#     logging.warning('------Start get company id from PDF------')
#     companies_id = []
#     wb_obj = openpyxl.load_workbook(pdf_to_excel_path)
#     sheet_obj = wb_obj.active
#     max_row = sheet_obj.max_row
#     for i in range(250):
#         rnum = random.randint(2, max_row)
#         cell_obj = sheet_obj.cell(row = rnum, column = 3)
#         print('Row: ' + str(rnum) + ' Value is: ' + cell_obj.value)
#         logging.info('Row: ' + str(rnum) + ' Value is: ' + cell_obj.value)
#         companies_id.append(cell_obj.value)
#     # print(sheet_obj.cell(row = max_row, column = 3).value)      
#     return companies_id