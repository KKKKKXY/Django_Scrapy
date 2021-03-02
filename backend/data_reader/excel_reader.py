import openpyxl
import random
import logging

def get_cid_from_excel(path):
    print('------Start get company id from EXCEL------')
    logging.warning('------Start get company id from PDF------')
    companies_id = []
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    for i in range(10): 
        rnum = random.randint(4, max_row)
        cell_obj = sheet_obj.cell(row = rnum, column = 2)
        print('Row: ' + str(rnum) + ' Value is: ' + cell_obj.value)
        logging.info('Row: ' + str(rnum) + ' Value is: ' + cell_obj.value)
        companies_id.append(cell_obj.value)
    return companies_id