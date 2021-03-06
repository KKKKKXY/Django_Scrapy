import os, time, re, pickle, signal
import pytesseract
from PIL import Image
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium import webdriver
import scrapy
import json
from scrapy_app.items import ScrapyAppItem
from scrapy.spiders import CrawlSpider #, Rule
import openpyxl
import random

class DbdcrawlerSpider(CrawlSpider):
    name = 'dbdcrawler_Thai'

    # def __init__(self, *args, **kwargs):
    #     # We are going to pass these args from our django view.
    #     # To make everything dynamic, we need to override them inside __init__ method
    #     self.url = kwargs.get('url')
    #     self.domain = kwargs.get('domain')
    #     self.start_urls = [self.url]
    #     self.allowed_domains = [self.domain]


    headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_0_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'
    }
    allowed_domains = ["datawarehouse.dbd.go.th"]

    def start_requests(self):
        companies_id = self.random_company()
        for i in companies_id:
            url = 'https://datawarehouse.dbd.go.th/company/profile/%s/%s' %(i[3],i)
            yield scrapy.Request(url, self.parse)

    def getCookie(self):
        cookie_path = '/Users/mya/DBDCrawler/scrapy_app/scrapy_app/spiders/temp/cookie.json'
        if os.path.isfile(cookie_path):
            try:
                with open(cookie_path, 'rb') as f: 
                    cookies = pickle.load(f)
            except EOFError:
                cookies = None

        for i in cookies:
            if i['name']=='JSESSIONID':
                cookies= i['value']
                break
        print(cookies)
        return cookies

    # def writeJsonFile(self, data):
    #     filePath = '/Users/mya/DBDCrawler/scrapy_app/scrapy_app/spiders/temp/thaiVersion.json'
    #     a_file = open(filePath, "w", encoding='utf-8')
    #     line = json.dumps(data, ensure_ascii=False) + "\n"
    #     a_file.write(line)

    # def readLoadsFile(self):
    #     loadsfilePath = '/Users/mya/DBDCrawler/scrapy_app/scrapy_app/spiders/temp/thaiVersion.json'
    #     print('------------Target Company Information------------')
    #     loadsdata = json.load(open(loadsfilePath))
    #     print(loadsdata)
    #     # print(loadsdata[0]['company_name'])
    #     return loadsdata

    def random_company(self):
        companies_path = '/Users/mya/DBDCrawler/scrapy_app/scrapy_app/spiders/db/dbd_oct2020.xlsx'
        companies_id = []
        wb_obj = openpyxl.load_workbook(companies_path)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        for i in range(5): 
            rnum = random.randint(4, max_row + 1)
            cell_obj = sheet_obj.cell(row = rnum, column = 2)
            companies_id.append(cell_obj.value)        
        print(companies_id)
        return companies_id

    def parse(self, response):
        print('------------START SCRAPING------------')
        time.sleep(5)
        objective = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[3]/div[5]/div/p/text()').get()
        if objective == '-' or objective == None:
            objective = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[3]/div[3]/div/p/text()').get()

        director_list = []
        directors = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[2]/div/div/ol/li/text()').getall()
        for i in directors:
            director_list.append(i.strip())

        raw_bussiness_type = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[3]/div[4]/div/p/text()').get()
        try:
            raw_bussiness_type = raw_bussiness_type.strip()
        except:
            raw_bussiness_type = 'ERRRRRRRRRRRRRRRRRRRORRRRRRRRRRRRRRRRRRRRRR:' + response.url.split('/')[-1]

        if raw_bussiness_type == '-'  or raw_bussiness_type == None:
            raw_bussiness_type = response.xpath('/html/body/div/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[3]/div[2]/div/p/text()').get().strip()

        tel = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[1]/div[2]/table/tr[3]/td[2]/text()').get().strip()
        if tel == None:
            tel = '-'

        fax = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[1]/div[2]/table/tr[4]/td[2]/text()').get().strip()
        if fax == None:
            fax = '-'

        website = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[1]/div[2]/table/tr[5]/td[2]/text()').get().strip()
        if website == None:
            website = '-'

        email = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[1]/div[2]/table/tr[6]/td[2]/text()').get().strip()
        if email == None:
            email = '-'

        item = ScrapyAppItem()
        item['company_id']          = response.xpath('/html/body/div/div[4]/div[2]/div/div[2]/div[1]/div/div[1]/p/text()').get().strip()
        item['company_name']        = response.xpath('/html/body/div/div[4]/div[2]/div[1]/div[1]/h2/text()').get().strip()
        item['company_type']        = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/table/tr[1]/th[2]/text()').get().strip()
        item['status']              = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/table/tr[3]/td[2]/text()').get().strip()
        item['objective']           = objective.strip()
        item['directors']           = director_list
        item['bussiness_type']      = raw_bussiness_type
        item['address']             = response.xpath('/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[1]/div[1]/div[2]/table/tr[2]/td/text()').get().strip()   
        item['tel']                 = tel
        item['fax']                 = fax
        item['website']             = website
        item['email']               = email

        # print('------------Target Target Target------------')
        # # self.writeJsonFile(item)
        # # item = self.readLoadsFile()
        # print(item)

        return item


# class DbdcrawlerSpider(CrawlSpider):

#     rules = (
#         Rule(LinkExtractor(allow=r'Items/'), callback='parse_item', follow=True),
#     )

#     def parse_item(self, response):
#         item = {}
#         #item['domain_id'] = response.xpath('//input[@id="sid"]/@value').get()
#         #item['name'] = response.xpath('//div[@id="name"]').get()
#         #item['description'] = response.xpath('//div[@id="description"]').get()
#         return item